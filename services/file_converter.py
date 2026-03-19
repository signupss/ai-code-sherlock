"""
File Converter — converts any output file to AI-readable text.

Supported:
  Text:   .txt, .log, .md, .rst, .csv, .tsv, .json, .yaml, .xml, .html
  Code:   .py, .js, .ts, .java, .go, .rs, .cs, .cpp, .c, .h, .sql
  Data:   .xlsx, .xls, .parquet, .feather (with optional pandas)
  Models: .pkl, .joblib → metadata only
  Binary: .npy, .npz → shape/stats
  Other:  → hex preview + size info
"""
from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Optional


MAX_TEXT_CHARS = 6000       # Per file
MAX_TABLE_ROWS = 50         # For CSV/Excel preview
MAX_TABLE_COLS = 20


class FileConverter:

    def convert(self, file_path: str, max_chars: int = MAX_TEXT_CHARS) -> ConvertedFile:
        """
        Convert any file to an AI-readable text representation.
        """
        path = Path(file_path)
        if not path.exists():
            return ConvertedFile(
                path=file_path, content=f"[Файл не найден: {file_path}]",
                file_type="missing", size_bytes=0
            )

        size = path.stat().st_size
        ext  = path.suffix.lower()

        try:
            if ext in (".txt", ".log", ".md", ".rst", ".py", ".js", ".ts",
                       ".java", ".go", ".rs", ".cs", ".cpp", ".c", ".h",
                       ".sql", ".sh", ".bat", ".ps1", ".yaml", ".yml",
                       ".toml", ".ini", ".cfg", ".env"):
                content = self._read_text(path, max_chars)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="text", size_bytes=size)

            elif ext == ".json":
                content = self._read_json(path, max_chars)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="json", size_bytes=size)

            elif ext in (".csv", ".tsv"):
                content = self._read_csv(path, max_chars)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="csv", size_bytes=size)

            elif ext in (".xlsx", ".xls"):
                content = self._read_excel(path, max_chars)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="excel", size_bytes=size)

            elif ext in (".html", ".htm"):
                content = self._read_html(path, max_chars)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="html", size_bytes=size)

            elif ext == ".xml":
                content = self._read_text(path, max_chars)  # XML is text
                return ConvertedFile(path=file_path, content=content,
                                     file_type="xml", size_bytes=size)

            elif ext in (".npy", ".npz"):
                content = self._read_numpy(path)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="numpy", size_bytes=size)

            elif ext in (".pkl", ".joblib", ".pickle"):
                content = self._read_pickle(path)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="pickle", size_bytes=size)

            elif ext in (".parquet", ".feather"):
                content = self._read_parquet(path, max_chars)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="parquet", size_bytes=size)

            elif ext in (".pt", ".pth", ".h5", ".hdf5"):
                content = self._read_model_file(path)
                return ConvertedFile(path=file_path, content=content,
                                     file_type="model", size_bytes=size)

            else:
                # Try as text first, fallback to binary info
                try:
                    content = self._read_text(path, max_chars)
                    return ConvertedFile(path=file_path, content=content,
                                         file_type="text", size_bytes=size)
                except Exception:
                    content = self._read_binary_info(path)
                    return ConvertedFile(path=file_path, content=content,
                                         file_type="binary", size_bytes=size)

        except Exception as e:
            return ConvertedFile(
                path=file_path,
                content=f"[Ошибка конвертации {path.name}: {e}]",
                file_type="error", size_bytes=size
            )

    def convert_for_ai(self, file_path: str, max_chars: int = MAX_TEXT_CHARS) -> str:
        """Returns formatted string ready for AI context injection."""
        cf = self.convert(file_path, max_chars)
        name = Path(file_path).name
        size_str = self._human_size(cf.size_bytes)
        header = f"### Файл: `{name}` [{cf.file_type}, {size_str}]\n"
        return header + "```\n" + cf.content + "\n```\n"

    # ── Readers ───────────────────────────────────────────

    @staticmethod
    def _read_text(path: Path, max_chars: int) -> str:
        with open(path, encoding="utf-8", errors="replace") as f:
            text = f.read()
        if len(text) <= max_chars:
            return text
        # Keep first 40% + last 40%, note cut
        head = int(max_chars * 0.4)
        tail = int(max_chars * 0.4)
        cut  = len(text) - head - tail
        return (text[:head] + f"\n\n... [{cut} символов вырезано] ...\n\n" + text[-tail:])

    @staticmethod
    def _read_json(path: Path, max_chars: int) -> str:
        with open(path, encoding="utf-8", errors="replace") as f:
            raw = f.read()
        try:
            data = json.loads(raw)
            # Pretty format with indent
            pretty = json.dumps(data, ensure_ascii=False, indent=2, default=str)
            if len(pretty) <= max_chars:
                return pretty
            # Truncate preserving structure
            return pretty[:max_chars] + "\n... [обрезано]"
        except json.JSONDecodeError:
            return raw[:max_chars]

    @staticmethod
    def _read_csv(path: Path, max_chars: int) -> str:
        try:
            import csv
            rows = []
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i > MAX_TABLE_ROWS:
                        rows.append(f"... и ещё строк (файл обрезан до {MAX_TABLE_ROWS})")
                        break
                    rows.append(" | ".join(row[:MAX_TABLE_COLS]))
            result = "\n".join(rows)
            if len(result) > max_chars:
                result = result[:max_chars] + "\n... [обрезано]"
            return result
        except Exception as e:
            return f"[Ошибка чтения CSV: {e}]"

    @staticmethod
    def _read_excel(path: Path, max_chars: int) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames[:5]:  # max 5 sheets
                ws = wb[sheet_name]
                parts.append(f"=== Лист: {sheet_name} ===")
                rows_seen = 0
                for row in ws.iter_rows(values_only=True):
                    if rows_seen >= MAX_TABLE_ROWS:
                        parts.append(f"... ещё строк (показано {MAX_TABLE_ROWS})")
                        break
                    cells = [str(c) if c is not None else "" for c in row[:MAX_TABLE_COLS]]
                    parts.append(" | ".join(cells))
                    rows_seen += 1
            result = "\n".join(parts)
        except ImportError:
            # Fallback: try CSV-style read
            try:
                import csv
                rows = []
                with open(path, "rb") as f:
                    sample = f.read(4096)
                result = f"[Excel файл, {len(sample)} байт. Установи openpyxl для полного чтения]"
            except Exception:
                result = "[Excel файл — openpyxl не установлен]"
        except Exception as e:
            result = f"[Ошибка чтения Excel: {e}]"

        if len(result) > max_chars:
            result = result[:max_chars] + "\n... [обрезано]"
        return result

    @staticmethod
    def _read_html(path: Path, max_chars: int) -> str:
        with open(path, encoding="utf-8", errors="replace") as f:
            html = f.read()
        # Strip tags
        try:
            import html as html_module
            text = re.sub(r"<[^>]+>", " ", html)
            text = html_module.unescape(text)
            text = re.sub(r"\s+", " ", text).strip()
            return text[:max_chars]
        except Exception:
            return html[:max_chars]

    @staticmethod
    def _read_numpy(path: Path) -> str:
        try:
            import numpy as np
            if path.suffix == ".npz":
                data = np.load(path)
                parts = [f"NPZ файл с ключами: {list(data.keys())}"]
                for k in list(data.keys())[:5]:
                    arr = data[k]
                    parts.append(f"  {k}: shape={arr.shape}, dtype={arr.dtype}, "
                                 f"min={arr.min():.4f}, max={arr.max():.4f}, mean={arr.mean():.4f}")
                return "\n".join(parts)
            else:
                arr = np.load(path)
                return (f"NumPy array: shape={arr.shape}, dtype={arr.dtype}\n"
                        f"min={arr.min():.4f}, max={arr.max():.4f}, mean={arr.mean():.4f}\n"
                        f"Первые значения: {arr.flat[:20].tolist()}")
        except ImportError:
            return f"[NumPy файл — numpy не установлен]"
        except Exception as e:
            return f"[Ошибка чтения NumPy: {e}]"

    @staticmethod
    def _read_pickle(path: Path) -> str:
        try:
            import pickle
            with open(path, "rb") as f:
                obj = pickle.load(f)
            type_name = type(obj).__name__
            info = [f"Pickle объект типа: {type_name}"]
            # Try to get useful info based on type
            if hasattr(obj, "__len__"):
                info.append(f"Длина: {len(obj)}")
            if hasattr(obj, "shape"):
                info.append(f"Shape: {obj.shape}")
            if hasattr(obj, "__dict__"):
                keys = list(obj.__dict__.keys())[:20]
                info.append(f"Атрибуты: {keys}")
            if hasattr(obj, "get_params"):
                # sklearn model
                try:
                    info.append(f"Параметры модели: {obj.get_params()}")
                except Exception:
                    pass
            return "\n".join(info)
        except Exception as e:
            return f"[Ошибка чтения Pickle: {e}]"

    @staticmethod
    def _read_parquet(path: Path, max_chars: int) -> str:
        try:
            import pandas as pd
            df = pd.read_parquet(path)
            info_lines = [
                f"Parquet: {df.shape[0]} строк × {df.shape[1]} столбцов",
                f"Столбцы: {list(df.columns[:20])}",
                f"Типы: {df.dtypes.to_dict()}",
                "",
                "Первые строки:",
                df.head(10).to_string(),
                "",
                "Статистика:",
                df.describe().to_string(),
            ]
            result = "\n".join(info_lines)
            return result[:max_chars]
        except ImportError:
            return "[Parquet файл — pandas не установлен]"
        except Exception as e:
            return f"[Ошибка чтения Parquet: {e}]"

    @staticmethod
    def _read_model_file(path: Path) -> str:
        ext = path.suffix.lower()
        size = path.stat().st_size
        info = [f"Файл модели: {path.name} ({FileConverter._human_size(size)})"]

        if ext in (".pt", ".pth"):
            try:
                import torch
                data = torch.load(path, map_location="cpu")
                if isinstance(data, dict):
                    info.append(f"Ключи state_dict: {list(data.keys())[:10]}")
                elif hasattr(data, "state_dict"):
                    info.append(f"PyTorch модель, параметров: {sum(p.numel() for p in data.parameters())}")
            except ImportError:
                info.append("[PyTorch не установлен]")
            except Exception as e:
                info.append(f"[Ошибка: {e}]")

        elif ext in (".h5", ".hdf5"):
            try:
                import h5py
                with h5py.File(path, "r") as f:
                    info.append(f"HDF5 ключи: {list(f.keys())[:10]}")
            except ImportError:
                info.append("[h5py не установлен]")
            except Exception as e:
                info.append(f"[Ошибка: {e}]")

        return "\n".join(info)

    @staticmethod
    def _read_binary_info(path: Path) -> str:
        size = path.stat().st_size
        with open(path, "rb") as f:
            header = f.read(64)
        hex_preview = " ".join(f"{b:02x}" for b in header)
        return (f"Бинарный файл: {path.name}\n"
                f"Размер: {FileConverter._human_size(size)}\n"
                f"Первые байты: {hex_preview}")

    @staticmethod
    def _human_size(size: int) -> str:
        for unit in ("Б", "КБ", "МБ", "ГБ"):
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} ТБ"


class ConvertedFile:
    def __init__(self, path: str, content: str, file_type: str, size_bytes: int):
        self.path = path
        self.content = content
        self.file_type = file_type
        self.size_bytes = size_bytes
        self.name = Path(path).name
